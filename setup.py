import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
import calendar
import os
from dotenv import load_dotenv
from crewai import Agent, Task, Crew, Process
from langchain_openai import ChatOpenAI
from crewai.tools import tool

# Carregar variáveis de ambiente
load_dotenv()
openai_api_key = os.getenv("OPENAI_API_KEY")
if not openai_api_key:
    raise ValueError("❌ Chave da OpenAI não encontrada. Verifique o arquivo .env.")

# Configurar LLM
llm = ChatOpenAI(
    model="gpt-4-turbo",
    temperature=0.1,
    api_key=openai_api_key
)

print("✅ Chave da OpenAI carregada com sucesso.")

# TOOLS
@tool
def load_and_consolidate_data():
    """
    Carrega todas as planilhas e consolida em um DataFrame único.
    Retorna o DataFrame consolidado em formato JSON para o agente.
    """
    files = {
        'admissao': 'ADMISSÃO ABRIL.xlsx',
        'afastamentos': 'AFASTAMENTOS.xlsx',
        'aprendiz': 'APRENDIZ.xlsx',
        'ativos': 'ATIVOS.xlsx',
        'dias_uteis': 'Base dias uteis.xlsx',
        'ferias': 'FÉRIAS.xlsx',
        'sindicato_valor': 'Base sindicato x valor.xlsx',
        'desligados': 'DESLIGADOS.xlsx',
        'estagio': 'ESTÁGIO.xlsx',
        'exterior': 'EXTERIOR.xlsx'
    }

    # Carregar dados
    data_dict = {}
    for key, filename in files.items():
        try:
            df = pd.read_excel(filename, sheet_name=0)
            data_dict[key] = df
            print(f"{filename} carregado: {len(df)} registros")
        except FileNotFoundError:
            print(f"Arquivo {filename} não encontrado.")
            data_dict[key] = pd.DataFrame()
        except Exception as e:
            print(f"Erro ao carregar {filename}: {e}")
            data_dict[key] = pd.DataFrame()

    # Consolidar dados
    if data_dict['ativos'].empty:
        raise ValueError("A base de ATIVOS está vazia.")
    
    df = data_dict['ativos'].copy()
    
    # Adicionar admissões
    if not data_dict['admissao'].empty and 'MATRICULA' in data_dict['admissao'].columns:
        admissao_clean = data_dict['admissao'][['MATRICULA', 'Admissão']].copy()
        admissao_clean.columns = ['MATRICULA', 'DATA_ADMISSAO']
        df = df.merge(admissao_clean, on='MATRICULA', how='left')

    # Adicionar afastamentos
    if not data_dict['afastamentos'].empty and 'MATRICULA' in data_dict['afastamentos'].columns:
        if 'DESC. SITUACAO' in data_dict['afastamentos'].columns:
            afastamento_clean = data_dict['afastamentos'][['MATRICULA', 'DESC. SITUACAO']].copy()
            afastamento_clean.columns = ['MATRICULA', 'SITUACAO_AFASTAMENTO']
            df = df.merge(afastamento_clean, on='MATRICULA', how='left')

    # Marcar aprendizes
    if not data_dict['aprendiz'].empty and 'MATRICULA' in data_dict['aprendiz'].columns:
        aprendiz_list = data_dict['aprendiz']['MATRICULA'].dropna().astype(int).tolist()
        df['E_APRENDIZ'] = df['MATRICULA'].isin(aprendiz_list)
    else:
        df['E_APRENDIZ'] = False

    # Marcar estagiários
    if not data_dict['estagio'].empty and 'MATRICULA' in data_dict['estagio'].columns:
        estagio_list = data_dict['estagio']['MATRICULA'].dropna().astype(int).tolist()
        df['E_ESTAGIARIO'] = df['MATRICULA'].isin(estagio_list)
    else:
        df['E_ESTAGIARIO'] = False

    # Adicionar desligados
    if not data_dict['desligados'].empty:
        desligados_df = data_dict['desligados'].copy()
        if 'MATRICULA ' in desligados_df.columns:
            desligados_df = desligados_df.rename(columns={'MATRICULA ': 'MATRICULA'})
        if 'DATA DEMISSÃO' in desligados_df.columns:
            desligados_df = desligados_df.rename(columns={'DATA DEMISSÃO': 'DATA_DESLIGAMENTO'})
        
        cols_to_merge = ['MATRICULA']
        if 'DATA_DESLIGAMENTO' in desligados_df.columns:
            cols_to_merge.append('DATA_DESLIGAMENTO')
        if 'COMUNICADO DE DESLIGAMENTO' in desligados_df.columns:
            cols_to_merge.append('COMUNICADO DE DESLIGAMENTO')
        
        df = df.merge(desligados_df[cols_to_merge], on='MATRICULA', how='left')

    # Adicionar férias
    if not data_dict['ferias'].empty and 'MATRICULA' in data_dict['ferias'].columns:
        if 'DIAS DE FÉRIAS' in data_dict['ferias'].columns:
            ferias_clean = data_dict['ferias'][['MATRICULA', 'DIAS DE FÉRIAS']].copy()
            df = df.merge(ferias_clean, on='MATRICULA', how='left')
        if 'DESC. SITUACAO' in data_dict['ferias'].columns:
            ferias_sit = data_dict['ferias'][['MATRICULA', 'DESC. SITUACAO']].copy()
            ferias_sit.columns = ['MATRICULA', 'SITUACAO_FERIAS']
            df = df.merge(ferias_sit, on='MATRICULA', how='left')
            
            # Consolidar situações
            if 'DESC. SITUACAO' in df.columns:
                df['DESC. SITUACAO'] = df['DESC. SITUACAO'].fillna(df['SITUACAO_FERIAS'])
            else:
                df['DESC. SITUACAO'] = df['SITUACAO_FERIAS']
            
            df.drop(columns=['SITUACAO_FERIAS'], inplace=True, errors='ignore')

    # Garantir colunas obrigatórias
    for col in ['DATA_DESLIGAMENTO', 'COMUNICADO DE DESLIGAMENTO', 'DIAS DE FÉRIAS']:
        if col not in df.columns:
            df[col] = pd.NaT if 'DATA' in col else ''

    # Salvar dados consolidados globalmente para outras ferramentas
    global consolidated_df, sindicato_valor_df
    consolidated_df = df
    sindicato_valor_df = data_dict['sindicato_valor']
    
    print(f"Dados consolidados: {len(df)} registros")
    return f"Dados consolidados com sucesso. Total de {len(df)} registros carregados."

@tool
def apply_exclusion_rules():
    """
    Aplica regras de exclusão aos dados consolidados.
    """
    global consolidated_df
    
    if consolidated_df is None:
        return "Erro: Dados não consolidados. Execute primeiro a consolidação."
    
    df = consolidated_df.copy()
    df['MOTIVO_EXCLUSAO'] = ''

    # Estagiários
    estagio_mask = df['E_ESTAGIARIO'] | df['TITULO DO CARGO'].str.contains('ESTAGIARIO', na=False, case=False)
    df.loc[estagio_mask, 'MOTIVO_EXCLUSAO'] = 'Estagiário'

    # Aprendizes
    aprendiz_mask = df['E_APRENDIZ']
    df.loc[aprendiz_mask, 'MOTIVO_EXCLUSAO'] = df.loc[aprendiz_mask, 'MOTIVO_EXCLUSAO'].apply(
        lambda x: (x + '; ' if x else '') + 'Aprendiz'
    )

    # Diretores
    diretor_mask = df['TITULO DO CARGO'].str.contains('DIRETOR', na=False, case=False)
    df.loc[diretor_mask, 'MOTIVO_EXCLUSAO'] = 'Diretor'

    # Afastados
    afastado_mask = (
        df['DESC. SITUACAO'].isin(['Licença Maternidade', 'Auxílio Doença']) |
        df.get('SITUACAO_AFASTAMENTO', pd.Series(dtype='object')).isin(['Licença Maternidade', 'Auxílio Doença'])
    )
    df.loc[afastado_mask, 'MOTIVO_EXCLUSAO'] = df.loc[afastado_mask, 'MOTIVO_EXCLUSAO'].apply(
        lambda x: (x + '; ' if x else '') + 'Afastado'
    )

    # Férias
    ferias_mask = df['DESC. SITUACAO'] == 'Férias'
    df.loc[ferias_mask, 'MOTIVO_EXCLUSAO'] = df.loc[ferias_mask, 'MOTIVO_EXCLUSAO'].apply(
        lambda x: (x + '; ' if x else '') + 'Férias'
    )

    # Filtrar elegíveis
    elegiveis_df = df[df['MOTIVO_EXCLUSAO'] == ''].copy()
    
    # Salvar dados filtrados globalmente
    global filtered_df
    filtered_df = elegiveis_df
    
    total_excluidos = len(df) - len(elegiveis_df)
    print(f"Exclusões aplicadas: {total_excluidos} excluídos, {len(elegiveis_df)} elegíveis")
    
    return f"Regras de exclusão aplicadas. {len(elegiveis_df)} colaboradores elegíveis ao VR de {len(df)} totais."

@tool
def calculate_vr_benefits():
    """
    Calcula os benefícios de VR para colaboradores elegíveis.
    """
    global filtered_df, sindicato_valor_df
    
    if filtered_df is None:
        return "Erro: Dados filtrados não disponíveis. Execute primeiro a filtragem."
    
    def get_vr_value_by_union(sindicato):
        """Retorna o valor do VR por dia com base no sindicato."""
        if sindicato_valor_df.empty or 'ESTADO' not in sindicato_valor_df.columns:
            return 35.00

        estado_map = {
            'SINDPPD RS': 'Rio Grande do Sul',
            'SINDPD SP': 'São Paulo',
            'SITEPD PR': 'Paraná',
            'SINDPD RJ': 'Rio de Janeiro'
        }

        estado = 'Rio Grande do Sul'  
        for key, value in estado_map.items():
            if key in str(sindicato):
                estado = value
                break

        for _, row in sindicato_valor_df.iterrows():
            if 'ESTADO' in row:
                estado_cell = str(row['ESTADO']).strip()
                if estado in estado_cell:
                    valor_str = str(row.get('VALOR', '35.00'))
                    if isinstance(valor_str, str):
                        valor_str = valor_str.replace('R$', '').replace(',', '.').strip()
                        try:
                            return float(valor_str)
                        except:
                            continue
                    else:
                        try:
                            return float(valor_str)
                        except:
                            continue
        return 35.00

    def calculate_working_days(row, reference_month=5, reference_year=2025):
        """Calcula dias úteis considerando admissão, desligamento e regra do dia 15."""
        first_day = date(reference_year, reference_month, 1)
        last_day = date(reference_year, reference_month, calendar.monthrange(reference_year, reference_month)[1])

        # Data de início
        start_date = first_day
        if pd.notna(row.get('DATA_ADMISSAO')):
            try:
                if isinstance(row['DATA_ADMISSAO'], str):
                    data_admissao = pd.to_datetime(row['DATA_ADMISSAO'], dayfirst=True).date()
                else:
                    data_admissao = row['DATA_ADMISSAO'].date() if isinstance(row['DATA_ADMISSAO'], pd.Timestamp) else first_day
                start_date = max(first_day, data_admissao)
            except:
                start_date = first_day

        # Data de fim
        end_date = last_day
        if pd.notna(row.get('DATA_DESLIGAMENTO')):
            try:
                if isinstance(row['DATA_DESLIGAMENTO'], str):
                    data_desligamento = pd.to_datetime(row['DATA_DESLIGAMENTO'], dayfirst=True).date()
                else:
                    data_desligamento = row['DATA_DESLIGAMENTO'].date() if isinstance(row['DATA_DESLIGAMENTO'], pd.Timestamp) else last_day

                # Regra de desligamento
                if row.get('COMUNICADO DE DESLIGAMENTO') == 'OK':
                    if data_desligamento.day <= 15:
                        return 0  # Não paga VR
                    else:
                        end_date = min(last_day, data_desligamento)
                else:
                    end_date = min(last_day, data_desligamento)
            except:
                end_date = last_day

        # Contar dias úteis (segunda a sexta)
        working_days = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:  # 0-4 = segunda a sexta
                working_days += 1
            current_date += timedelta(days=1)

        return working_days

    # Aplicar cálculos
    df = filtered_df.copy()
    df['DIAS_UTEIS'] = df.apply(lambda row: calculate_working_days(row), axis=1)
    df['VR_POR_DIA'] = df['Sindicato'].apply(get_vr_value_by_union)
    df['VR_TOTAL'] = df['DIAS_UTEIS'] * df['VR_POR_DIA']
    df['VR_EMPRESA'] = df['VR_TOTAL'] * 0.8
    df['VR_PROFISSIONAL'] = df['VR_TOTAL'] * 0.2

    # Salvar dados calculados globalmente
    global calculated_df
    calculated_df = df
    
    print(f"Cálculos de VR concluídos para {len(df)} colaboradores")
    return f"Cálculos de VR concluídos. {len(df)} colaboradores processados."

@tool
def generate_final_spreadsheet():
    """
    Gera a planilha final e executa validações.
    """
    global calculated_df
    
    if calculated_df is None:
        return "Erro: Dados calculados não disponíveis. Execute primeiro os cálculos."
    
    # Definir arquivo de saída
    output_filename = 'VR_MENSAL_05.2025_OUTPUT.xlsx'
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, output_filename)
    
    # Colunas obrigatórias
    required_columns = [
        'MATRICULA', 'EMPRESA', 'TITULO DO CARGO', 'Sindicato',
        'DIAS_UTEIS', 'VR_POR_DIA', 'VR_TOTAL', 'VR_EMPRESA', 'VR_PROFISSIONAL'
    ]
    
    # Garantir que todas as colunas existem
    df = calculated_df.copy()
    for col in required_columns:
        if col not in df.columns:
            df[col] = ''
    
    # Preparar DataFrame final
    output_df = df[required_columns].copy()
    output_df.columns = [
        'Matricula', 'Empresa', 'Cargo', 'Sindicato',
        'Dias Uteis', 'Valor VR/Dia', 'Valor Total VR', 
        'Valor Empresa (80%)', 'Valor Profissional (20%)'
    ]
    
    # Salvar planilha com formatação adequada
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            output_df.to_excel(writer, index=False, sheet_name='VR Mensal 05.2025')
            
            # Acessar a planilha para formatação
            worksheet = writer.sheets['VR Mensal 05.2025']
            
            # Ajustar largura das colunas
            column_widths = {
                'A': 12,  # Matricula
                'B': 12,  # Admissão
                'C': 40,  # Sindicato do Colaborador
                'D': 12,  # Competência
                'E': 8,   # Dias
                'F': 15,  # VALOR DIÁRIO VR
                'G': 12,  # TOTAL
                'H': 15,  # Custo empresa
                'I': 18   # Desconto profissional
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
        
        print(f"Planilha gerada com sucesso em: {output_path}")
        
    except Exception as e:
        return f"Erro ao salvar planilha: {e}"
    
    # Executar validações 
    errors = []
    
    if (output_df['Valor Total VR'] < 0).any():
        errors.append("Valores de VR negativos encontrados.")
    
    if (output_df['Dias Uteis'] < 0).any():
        errors.append("Dias úteis negativos encontrados.")
    
    if output_df['Matricula'].duplicated().any():
        errors.append("Matrículas duplicadas encontradas.")
    
    zero_vr = output_df[(output_df['Dias Uteis'] > 0) & (output_df['Valor Total VR'] == 0)]
    if not zero_vr.empty:
        errors.append(f"{len(zero_vr)} registros com dias úteis > 0 mas VR = 0.")
    
    # Calcular totais para resumo
    total_vr = output_df['Valor Total VR'].sum()
    total_empresa = output_df['Valor Empresa (80%)'].sum()
    total_profissional = output_df['Valor Profissional (20%)'].sum()
    
    # Salvar informações globalmente para o resumo executivo
    global summary_data
    summary_data = {
        'registros_processados': len(output_df),
        'total_vr': total_vr,
        'custo_empresa': total_empresa,
        'desconto_profissional': total_profissional,
        'validacao_status': 'SUCESSO' if len(errors) == 0 else 'FALHA',
        'validacao_msg': "🎉 Validação bem-sucedida! A planilha está pronta para envio." if len(errors) == 0 else "⚠️ Erros encontrados na validação",
        'errors': errors
    }
    
    # Preparar resultado
    if len(errors) == 0:
        validation_msg = "🎉 Validação bem-sucedida! A planilha está pronta para envio."
        status = "SUCESSO"
    else:
        validation_msg = "⚠️  Erros encontrados na validação:\n" + "\n".join([f"  - {e}" for e in errors])
        status = "FALHA"
    
    result_msg = f"""
                    📁 PLANILHA GERADA: {output_path}
                    📊 REGISTROS PROCESSADOS: {len(output_df)}
                    💰 TOTAL VR: R$ {total_vr:,.2f}
                    🏢 CUSTO EMPRESA (80%): R$ {total_empresa:,.2f}
                    👤 DESCONTO PROFISSIONAL (20%): R$ {total_profissional:,.2f}
                    📋 VALIDAÇÃO: {validation_msg}
                    🏁 STATUS: {status}

                    📋 AMOSTRA DOS DADOS GERADOS:
                    {output_df.head(3).to_string(index=False)}
    """
    
    print(result_msg)
    return result_msg

@tool
def generate_executive_summary():
    """
    Gera uma planilha de resumo executivo com as informações consolidadas do processo.
    """
    global summary_data
    
    if summary_data is None:
        return "Erro: Dados de resumo não disponíveis. Execute primeiro a geração da planilha principal."
    
    # Criar DataFrame com resumo executivo
    current_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    resumo_data = {
        'Métrica': [
            'Data/Hora da Execução',
            'Competência',
            'Registros Processados',
            'Total Vale Refeição (R$)',
            'Custo Empresa - 80% (R$)',
            'Desconto Profissional - 20% (R$)',
            'Status da Validação',
            'Observações'
        ],
        'Valor': [
            current_date,
            'Maio/2025',
            f"{summary_data['registros_processados']:,}".replace(',', '.'),
            f"R$ {summary_data['total_vr']:,.2f}".replace(',', '.'),
            f"R$ {summary_data['custo_empresa']:,.2f}".replace(',', '.'),
            f"R$ {summary_data['desconto_profissional']:,.2f}".replace(',', '.'),
            summary_data['validacao_status'],
            summary_data['validacao_msg']
        ]
    }
    
    resumo_df = pd.DataFrame(resumo_data)
    
    # Definir arquivo de saída para o resumo
    summary_filename = 'RESUMO_EXECUTIVO_VR_05.2025.xlsx'
    script_dir = os.path.dirname(os.path.abspath(__file__))
    summary_path = os.path.join(script_dir, summary_filename)
    
    try:
        with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
            # Aba principal - Resumo Executivo
            resumo_df.to_excel(writer, index=False, sheet_name='Resumo Executivo')
            
            # Aba adicional - Detalhes de Validação (se houver erros)
            if summary_data['errors']:
                validacao_data = {
                    'Tipo de Erro': summary_data['errors'],
                    'Status': ['ATENÇÃO'] * len(summary_data['errors'])
                }
                validacao_df = pd.DataFrame(validacao_data)
                validacao_df.to_excel(writer, index=False, sheet_name='Erros de Validação')
            
            # Formatação da aba principal
            worksheet = writer.sheets['Resumo Executivo']
            
            # Ajustar largura das colunas
            worksheet.column_dimensions['A'].width = 30  # Métrica
            worksheet.column_dimensions['B'].width = 40  # Valor
            
            # Formatação dos cabeçalhos
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:  # Primeira linha (cabeçalhos)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Formatação das células de valor monetário
            currency_rows = [4, 5, 6]  # Linhas com valores monetários
            for row_num in currency_rows:
                cell = worksheet[f'B{row_num}']
                cell.alignment = Alignment(horizontal="right")
            
            # Formatação da célula de status
            status_cell = worksheet['B7']
            if summary_data['validacao_status'] == 'SUCESSO':
                status_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
        
        print(f"📊 Resumo Executivo gerado com sucesso em: {summary_path}")
        
        result_msg = f"""
        🎯 RESUMO EXECUTIVO GERADO COM SUCESSO!
        
        📁 Arquivo: {summary_filename}
        📍 Local: {summary_path}
        
        📊 DADOS CONSOLIDADOS:
        ✅ Registros Processados: {summary_data['registros_processados']:,}
        💰 Total VR: R$ {summary_data['total_vr']:,.2f}
        🏢 Custo Empresa (80%): R$ {summary_data['custo_empresa']:,.2f}
        👤 Desconto Profissional (20%): R$ {summary_data['desconto_profissional']:,.2f}
        📋 Validação: {summary_data['validacao_msg']}
        🏁 Status: {summary_data['validacao_status']}
        
        📈 O resumo executivo está pronto para apresentação à gestão!
        """
        
        return result_msg
        
    except Exception as e:
        return f"❌ Erro ao gerar resumo executivo: {e}"

# VARIÁVEIS GLOBAIS PARA COMPARTILHAR DADOS
consolidated_df = None
filtered_df = None
calculated_df = None
sindicato_valor_df = None
summary_data = None

# DEFINIÇÃO DOS AGENTES

data_consolidator = Agent(
    role='Especialista em Consolidação de Dados',
    goal='Carregar, mesclar e limpar as bases de dados em uma única base unificada.',
    backstory="""Você é um analista de dados sênior com expertise em integração de múltiplas fontes.
    Sua missão é garantir que todas as informações de admissão, desligamento, férias, sindicatos e exclusões
    estejam corretamente associadas por matrícula para o cálculo do VR.""",
    verbose=True,
    allow_delegation=False,
    tools=[load_and_consolidate_data],
    llm=llm
)

filter_agent = Agent(
    role='Especialista em Regras de Exclusão e Elegibilidade',
    goal='Aplicar todas as regras de exclusão para filtrar colaboradores inelegíveis ao VR.',
    backstory="""Você é um especialista em conformidade de RH. Sabe exatamente quais colaboradores
    devem ser excluídos da tabela: estagiários, aprendizes, diretores, afastados, em férias ou no exterior.
    Garanta que apenas colaboradores elegíveis avancem no processo.""",
    verbose=True,
    allow_delegation=False,
    tools=[apply_exclusion_rules],
    llm=llm
)

calculator_agent = Agent(
    role='Analista de Benefícios e Cálculos Proporcionais',
    goal='Calcular o número exato de dias úteis e o valor proporcional de VR para cada colaborador elegível.',
    backstory="""Você é um analista financeiro especializado em cálculos proporcionais de benefícios.
    Domina as regras de sindicatos, admissões, desligamentos (com regra do dia 15, se desligamento antes do dia 15, 
    não tem direito ao vale, se o desligamento for após o dia 15, tem direito ao vale de forma proporcional) e férias.
    Sua precisão é crítica para evitar erros no pagamento.""",
    verbose=True,
    allow_delegation=False,
    tools=[calculate_vr_benefits],
    llm=llm
)

validator_agent = Agent(
    role='Especialista em Qualidade e Geração de Layout',
    goal='Gerar a planilha final no formato exigido e executar validações para garantir a qualidade dos dados.',
    backstory="""Você é um QA sênior focado em processos financeiros. Sua responsabilidade é entregar
    uma planilha impecável, no formato exato solicitado, livre de erros e pronta para envio ao fornecedor.
    Você segue checklists rigorosos e documenta qualquer anomalia.""",
    verbose=True,
    allow_delegation=False,
    tools=[generate_final_spreadsheet],
    llm=llm
)

executive_summary_agent = Agent(
    role='Analista de Resumo Executivo e Relatórios Gerenciais',
    goal='Gerar resumos executivos profissionais com métricas consolidadas para apresentação à gestão.',
    backstory="""Você é um analista de business intelligence especializado em criar relatórios executivos
    claros e objetivos. Sua expertise está em transformar dados operacionais em insights estratégicos
    para tomada de decisão da alta gestão. Você cria planilhas formatadas profissionalmente com
    métricas-chave, status de validação e informações relevantes para o acompanhamento de processos.""",
    verbose=True,
    allow_delegation=False,
    tools=[generate_executive_summary],
    llm=llm
)

# DEFINIÇÃO DAS TAREFAS

task1 = Task(
    description="""Execute a ferramenta de consolidação de dados para carregar todas as planilhas:
    ATIVOS, ADMISSÃO ABRIL, AFASTAMENTOS, APRENDIZ, Base dias uteis, FÉRIAS, Base sindicato x valor, 
    DESLIGADOS, ESTÁGIO, EXTERIOR.
    
    Use a ferramenta load_and_consolidate_data() para:
    1. Carregar todas as planilhas
    2. Consolidar em um único DataFrame usando MATRICULA como chave
    3. Garantir que todas as colunas necessárias estejam presentes
    
    Relate o resultado da consolidação.""",
    expected_output="Relatório de consolidação com número de registros carregados e status.",
    agent=data_consolidator
)

task2 = Task(
    description="""Execute a ferramenta de aplicação de regras de exclusão.
    
    Use a ferramenta apply_exclusion_rules() para remover da planilha consolidada:
    - Estagiários (base ESTÁGIO e cargo contendo 'ESTAGIARIO')
    - Aprendizes (base APRENDIZ)
    - Diretores (cargo contendo 'DIRETOR')
    - Afastados (Licença Maternidade, Auxílio Doença)
    - Colaboradores em férias (DESC. SITUACAO = 'Férias')
    
    Relate quantos colaboradores foram excluídos e quantos são elegíveis.""",
    expected_output="Relatório de exclusões com números de excluídos por categoria e total de elegíveis.",
    agent=filter_agent
)

task3 = Task(
    description="""Execute a ferramenta de cálculo de benefícios de VR.
    
    Use a ferramenta calculate_vr_benefits() para calcular:
    1. Dias úteis em maio/2025 considerando admissão, desligamento e regra do dia 15
    2. Valor diário de VR baseado no sindicato
    3. Valor total de VR (dias úteis × valor diário)
    4. Valor empresa (80%) e valor profissional (20%)
    5. Regra do dia 15, se desligamento ocorreu antes do dia 15, não tem direito ao vale, se for após o dia 15, tem direito ao vale de forma proporcional
    
    Relate o resultado dos cálculos.""",
    expected_output="Relatório de cálculos com número de colaboradores processados e valores calculados.",
    agent=calculator_agent
)

task4 = Task(
    description="""Execute a ferramenta de geração da planilha final.
    
    Use a ferramenta generate_final_spreadsheet() para:
    1. Gerar a planilha 'VR_MENSAL_05.2025_OUTPUT.xlsx' com o formato correto
    2. Executar todas as validações de qualidade
    3. Fornecer relatório detalhado de status
    
    A planilha deve conter as colunas:
    'Matrícula', 'Empresa', 'Cargo', 'Sindicato', 'Dias Úteis', 'Valor VR/Dia', 
    'Valor Total VR', 'Valor Empresa (80%)', 'Valor Profissional (20%)'
    
    Relate onde a planilha foi salva e o resultado das validações.""",
    expected_output="Relatório final com localização da planilha, validações e status de conclusão.",
    agent=validator_agent
)

task5 = Task(
    description="""Execute a ferramenta de geração do resumo executivo.
    
    Use a ferramenta generate_executive_summary() para:
    1. Criar a planilha 'RESUMO_EXECUTIVO_VR_05.2025.xlsx'
    2. Incluir métricas consolidadas do processo
    3. Apresentar dados em formato executivo para gestão
    
    O resumo deve conter:
    - Data/Hora da execução
    - Competência (Maio/2025)
    - Total de registros processados
    - Valor total do Vale Refeição
    - Custo para a empresa (80%)
    - Desconto do profissional (20%)
    - Status da validação
    - Observações relevantes
    
    Relate onde o resumo executivo foi salvo.""",
    expected_output="Relatório com localização da planilha de resumo executivo e métricas consolidadas.",
    agent=executive_summary_agent
)

# CRIAÇÃO DO FLUXO DE EXECUÇÃO

crew = Crew(
    agents=[data_consolidator, filter_agent, calculator_agent, validator_agent, executive_summary_agent],
    tasks=[task1, task2, task3, task4, task5],
    verbose=True,
    process=Process.sequential
)

# EXECUÇÃO PRINCIPAL

if __name__ == "__main__":
    print("Iniciando o processo de automação do calculo do VR...")
    print("=" * 60)
    
    try:
        result = crew.kickoff()
        print("\n" + "=" * 60)
        print("✅ PROCESSO CONCLUÍDO COM SUCESSO!")
        print("=" * 60)
        print("Resultado final:", result)
        
    except Exception as e:
        print(f"\n❌ ERRO DURANTE A EXECUÇÃO: {e}")
        print("Verifique se todas as planilhas estão na pasta correta.")